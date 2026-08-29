"""Spreadsheet import: parse -> map -> validate -> preview -> commit.

The flow is two calls on purpose. ``preview`` parses the file, infers a column
mapping, validates every row and stores the verdicts; ``commit`` writes the
approved subset. Nothing reaches the catalog until the vendor has seen what
will land, and because the parsed rows are persisted between the two calls,
what they approved is exactly what commits -- the file is never re-read.

Partial import is the default. One bad row in a two-hundred-row price list
should not cost the vendor the other hundred and ninety-nine; the bad rows come
back with per-field messages so they can be fixed and re-uploaded.

Committing goes through ``CsvSource``, the same CatalogSource adapter every
other route uses, so an imported row is upserted by exactly the rules a
manually typed or API-synced row is. Adding XLSX support here changed no
endpoint and no agent code, which is the point of the adapter.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Sequence
from uuid import UUID

import structlog

from app.core.config import settings
from app.schemas.imports import ImportTemplate, ImportTemplateColumn
from app.services.catalog_sources import SourceItem

log = structlog.get_logger(__name__)


# --------------------------------------------------------------------------
# The target schema
#
# One table drives the downloadable template, the mapping dropdowns and the
# validator, so those three can never drift apart.
# --------------------------------------------------------------------------
@dataclass(frozen=True, slots=True)
class TargetField:
    name: str
    required: bool
    kind: str  # text | money | int
    example: str
    note: str | None = None
    #: Header spellings that map here without the vendor touching anything.
    aliases: tuple[str, ...] = ()


TARGET_FIELDS: tuple[TargetField, ...] = (
    TargetField(
        "sku", True, "text", "LAT-5550",
        "Unique per vendor. Re-importing the same SKU updates that item.",
        ("sku", "code", "item code", "product code", "part number", "partno", "id"),
    ),
    TargetField(
        "title", True, "text", "Dell Latitude 5550 laptop", None,
        ("title", "name", "product", "product name", "item", "description short"),
    ),
    TargetField(
        "price", True, "money", "174000",
        "Digits only or with separators. Currency comes from your profile.",
        ("price", "unit price", "rate", "cost", "amount", "mrp"),
    ),
    TargetField(
        "stock", True, "int", "240", "Whole units on hand.",
        ("stock", "qty", "quantity", "on hand", "inventory", "available"),
    ),
    TargetField(
        "delivery_days", False, "int", "7",
        "Left blank, your vendor default applies. With neither, buyers see "
        "reduced data confidence and the item scores lower.",
        ("delivery days", "delivery", "lead time", "lead time days", "eta days"),
    ),
    TargetField(
        "warranty_months", False, "int", "24", "Same fallback as delivery.",
        ("warranty months", "warranty", "guarantee", "warranty period"),
    ),
    TargetField(
        "description", False, "text", "i7 - 16GB - 512GB", None,
        ("description", "details", "spec", "specification", "long description"),
    ),
    TargetField(
        "category", False, "text", "IT hardware", None,
        ("category", "type", "group", "department"),
    ),
    TargetField(
        "brand", False, "text", "Dell", None,
        ("brand", "make", "manufacturer", "vendor brand"),
    ),
    TargetField(
        "sale_price", False, "money", "165000",
        "Optional discounted price. Quotes use it when present.",
        ("sale price", "discount price", "offer price", "special price"),
    ),
)

_BY_NAME = {f.name: f for f in TARGET_FIELDS}
REQUIRED_FIELDS = tuple(f.name for f in TARGET_FIELDS if f.required)


def template() -> ImportTemplate:
    return ImportTemplate(
        columns=[
            ImportTemplateColumn(
                name=f.name, required=f.required, example=f.example, note=f.note
            )
            for f in TARGET_FIELDS
        ]
    )


def template_csv() -> str:
    """The downloadable starter file: header plus one worked example row."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow([f.name for f in TARGET_FIELDS])
    writer.writerow([f.example for f in TARGET_FIELDS])
    return buf.getvalue()


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------
class ImportError_(ValueError):
    """A whole-file problem: unreadable, empty, or no usable header."""


def _normalise(header: str) -> str:
    """Fold a header for comparison: case, punctuation and spacing."""
    return re.sub(r"[^a-z0-9]+", " ", header.strip().lower()).strip()


def suggest_mapping(columns: Sequence[str]) -> list[dict[str, str]]:
    """Guess the column mapping so the common file needs no manual work.

    Exact-normalised match first, then alias, then a contains-match as a last
    resort. A target is claimed once -- two columns cannot both map to price.
    """
    mapping: list[dict[str, str]] = []
    claimed: set[str] = set()

    def claim(column: str, field: str) -> None:
        mapping.append({"source_column": column, "target_field": field})
        claimed.add(field)

    normalised = {c: _normalise(c) for c in columns}

    for column, norm in normalised.items():
        if norm in _BY_NAME and norm not in claimed:
            claim(column, norm)

    for column, norm in normalised.items():
        if any(m["source_column"] == column for m in mapping):
            continue
        for field in TARGET_FIELDS:
            if field.name in claimed:
                continue
            if norm in field.aliases:
                claim(column, field.name)
                break

    for column, norm in normalised.items():
        if any(m["source_column"] == column for m in mapping) or not norm:
            continue
        for field in TARGET_FIELDS:
            if field.name in claimed:
                continue
            if any(a in norm or norm in a for a in field.aliases):
                claim(column, field.name)
                break

    return mapping


def parse_file(filename: str, content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Return (columns, rows-as-dicts). Raises ImportError_ on a bad file."""
    if len(content) > settings.import_max_file_bytes:
        raise ImportError_(
            f"File is larger than the "
            f"{settings.import_max_file_bytes // (1024 * 1024)} MB limit."
        )
    if not content.strip():
        raise ImportError_("The file is empty.")

    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _parse_excel(content)
    return _parse_csv(content)


def _decode(content: bytes) -> str:
    # Excel on Windows still writes cp1252 and a UTF-8 BOM is common; try the
    # likely encodings rather than failing a vendor's file on a stray byte.
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportError_("Could not read the file's text encoding.")


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = _decode(content)
    sample = text[:8192]
    try:
        dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # a single-column file sniffs as an error

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header = next(reader)
    except StopIteration as exc:
        raise ImportError_("The file has no header row.") from exc

    columns = _dedupe_headers(header)
    if not any(c.strip() for c in columns):
        raise ImportError_("The header row is blank.")

    rows: list[dict[str, str]] = []
    for values in reader:
        if not any((v or "").strip() for v in values):
            continue  # trailing blank lines are not data
        rows.append(_zip_row(columns, values))
    return columns, rows


def _parse_excel(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ImportError_("Excel support is unavailable on this server.") from exc

    try:
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises many types
        raise ImportError_("That file is not a readable .xlsx workbook.") from exc

    sheet = book.active
    if sheet is None:
        raise ImportError_("The workbook has no sheets.")

    iterator = sheet.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration as exc:
        raise ImportError_("The sheet is empty.") from exc

    columns = _dedupe_headers(["" if h is None else str(h) for h in header])
    rows: list[dict[str, str]] = []
    for values in iterator:
        cells = ["" if v is None else str(v).strip() for v in values]
        if not any(cells):
            continue
        rows.append(_zip_row(columns, cells))
    book.close()
    return columns, rows


def _dedupe_headers(header: Iterable[Any]) -> list[str]:
    """``price``, ``price`` -> ``price``, ``price (2)``.

    Duplicated headers are common in exported files and would otherwise
    silently overwrite each other in the row dict.
    """
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, raw in enumerate(header):
        name = ("" if raw is None else str(raw)).strip() or f"column {i + 1}"
        count = seen.get(name.lower(), 0) + 1
        seen[name.lower()] = count
        out.append(name if count == 1 else f"{name} ({count})")
    return out


def _zip_row(columns: Sequence[str], values: Sequence[Any]) -> dict[str, str]:
    row: dict[str, str] = {}
    for i, column in enumerate(columns):
        value = values[i] if i < len(values) else ""
        row[column] = "" if value is None else str(value).strip()
    return row


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------
_MONEY_STRIP = re.compile(r"[^0-9.\-]")


def _coerce(field: TargetField, raw: str) -> tuple[Any, str | None]:
    """Return (value, error). A blank optional field is (None, None)."""
    text = (raw or "").strip()
    if not text:
        if field.required:
            return None, "Required, but this cell is empty."
        return None, None

    if field.kind == "text":
        return text, None

    if field.kind == "money":
        cleaned = _MONEY_STRIP.sub("", text)
        if not cleaned or cleaned in {"-", ".", "-."}:
            return None, f"'{text}' is not a number."
        try:
            value = Decimal(cleaned)
        except InvalidOperation:
            return None, f"'{text}' is not a number."
        if value < 0:
            return None, "Cannot be negative."
        if value == 0:
            return None, "Must be greater than zero."
        return value, None

    # int
    cleaned = _MONEY_STRIP.sub("", text)
    if not cleaned:
        return None, f"'{text}' is not a whole number."
    try:
        value = Decimal(cleaned)
    except InvalidOperation:
        return None, f"'{text}' is not a whole number."
    if value != value.to_integral_value():
        return None, f"'{text}' must be a whole number."
    number = int(value)
    if number < 0:
        return None, "Cannot be negative."
    return number, None


def validate_rows(
    raw_rows: Sequence[dict[str, str]],
    mapping: Sequence[dict[str, str]],
    *,
    existing_skus: set[str],
) -> list[dict[str, Any]]:
    """Per-row verdicts. Never raises -- a bad row is data, not an exception."""
    by_target = {m["target_field"]: m["source_column"] for m in mapping}
    missing_required = [f for f in REQUIRED_FIELDS if f not in by_target]

    # A SKU repeated inside one file would upsert twice and the later row would
    # silently win. Flag it instead.
    seen_in_file: set[str] = set()
    verdicts: list[dict[str, Any]] = []

    for index, raw in enumerate(raw_rows, start=1):
        errors: list[dict[str, str]] = []
        parsed: dict[str, Any] = {}

        for field_name in missing_required:
            errors.append(
                {
                    "field": field_name,
                    "message": f"No column is mapped to {field_name}.",
                }
            )

        for target, column in by_target.items():
            field = _BY_NAME.get(target)
            if field is None:
                continue
            value, error = _coerce(field, raw.get(column, ""))
            if error:
                errors.append({"field": target, "message": error})
            elif value is not None:
                parsed[target] = value

        sku = str(parsed.get("sku", "")).strip()
        duplicate_in_file = bool(sku) and sku in seen_in_file
        if duplicate_in_file:
            errors.append(
                {
                    "field": "sku",
                    "message": f"'{sku}' appears more than once in this file.",
                }
            )
        if sku:
            seen_in_file.add(sku)

        if (
            parsed.get("sale_price") is not None
            and parsed.get("price") is not None
            and parsed["sale_price"] > parsed["price"]
        ):
            errors.append(
                {
                    "field": "sale_price",
                    "message": "Sale price is higher than the price.",
                }
            )

        missing_terms = [
            name
            for name in ("delivery_days", "warranty_months")
            if parsed.get(name) is None
        ]

        verdicts.append(
            {
                "row_number": index,
                "raw": raw,
                "parsed": _jsonable(parsed) if parsed else None,
                "errors": errors,
                "is_duplicate_sku": bool(sku)
                and not duplicate_in_file
                and sku in existing_skus,
                "missing_terms": missing_terms,
            }
        )
    return verdicts


def _jsonable(parsed: dict[str, Any]) -> dict[str, Any]:
    """JSONB cannot hold a Decimal, so money is stored as its exact string.

    Not as a float: everything downstream of here is NUMERIC, and a price that
    round-trips through binary floating point is the kind of thing that shows
    up as a one-rupee discrepancy on a purchase order months later.
    """
    return {
        k: (str(v) if isinstance(v, Decimal) else v) for k, v in parsed.items()
    }


def to_source_items(
    parsed_rows: Iterable[dict[str, Any]], *, currency: str
) -> list[SourceItem]:
    """Validated rows -> the adapter's input type."""
    items: list[SourceItem] = []
    for parsed in parsed_rows:
        items.append(
            SourceItem(
                sku=str(parsed["sku"]),
                title=str(parsed["title"]),
                price=Decimal(str(parsed["price"])),
                stock=int(parsed["stock"]),
                description=parsed.get("description"),
                category=parsed.get("category"),
                brand=parsed.get("brand"),
                sale_price=(
                    Decimal(str(parsed["sale_price"]))
                    if parsed.get("sale_price") is not None
                    else None
                ),
                delivery_days=parsed.get("delivery_days"),
                warranty_months=parsed.get("warranty_months"),
                currency=currency,
            )
        )
    return items


__all__ = [
    "TARGET_FIELDS",
    "REQUIRED_FIELDS",
    "ImportError_",
    "parse_file",
    "suggest_mapping",
    "template",
    "template_csv",
    "to_source_items",
    "validate_rows",
]
